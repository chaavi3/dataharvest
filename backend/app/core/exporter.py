"""Export extracted data to CSV, XLSX, HTML, PDF."""

import csv
import io
import logging
from pathlib import Path

from ..models.schema import ColumnDef

logger = logging.getLogger(__name__)


def export_csv(rows: list[dict], columns: list[ColumnDef]) -> bytes:
    buf = io.StringIO()
    buf.write("\ufeff")  # BOM for Excel
    col_names = [c.name for c in columns]
    writer = csv.DictWriter(buf, fieldnames=col_names, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in col_names})
    return buf.getvalue().encode("utf-8")


def export_xlsx(rows: list[dict], columns: list[ColumnDef]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    col_names = [c.name for c in columns]

    for ci, name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = Font(bold=True)

    for ri, row in enumerate(rows, 2):
        for ci, name in enumerate(col_names, 1):
            ws.cell(row=ri, column=ci, value=row.get(name, ""))

    for ci, name in enumerate(col_names, 1):
        max_len = max(len(str(name)), *(len(str(r.get(name, ""))) for r in rows)) if rows else len(str(name))
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_html(rows: list[dict], columns: list[ColumnDef], title: str = "DataHarvest Export") -> bytes:
    col_names = [c.name for c in columns]
    lines = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'>",
        f"<title>{title}</title>",
        "<style>",
        "body{font-family:system-ui,sans-serif;margin:2rem}",
        "table{border-collapse:collapse;width:100%}",
        "th,td{border:1px solid #ddd;padding:8px 12px;text-align:left}",
        "th{background:#f5f5f5;font-weight:600}",
        "tr:nth-child(even){background:#fafafa}",
        "</style></head><body>",
        f"<h1>{title}</h1>",
        "<table><thead><tr>",
    ]
    for n in col_names:
        lines.append(f"<th>{_esc(n)}</th>")
    lines.append("</tr></thead><tbody>")
    for row in rows:
        lines.append("<tr>")
        for n in col_names:
            lines.append(f"<td>{_esc(str(row.get(n, '')))}</td>")
        lines.append("</tr>")
    lines.append("</tbody></table></body></html>")
    return "\n".join(lines).encode("utf-8")


def export_pdf(rows: list[dict], columns: list[ColumnDef], title: str = "DataHarvest Export") -> bytes:
    html_bytes = export_html(rows, columns, title)
    try:
        from weasyprint import HTML
        return HTML(string=html_bytes.decode("utf-8")).write_pdf()
    except ImportError:
        logger.warning("weasyprint not installed; falling back to HTML-as-PDF")
        return html_bytes


def _esc(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
